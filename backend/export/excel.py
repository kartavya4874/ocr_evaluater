"""
Excel export module — generates course result spreadsheets with 4 sheets.
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Optional, Any
from pathlib import Path
import zipfile
import statistics

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
FLAG_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")


def _apply_header_style(ws, row, col_count):
    """Apply header style to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_cell_border(ws, row, col_count):
    """Apply border to cells in a row."""
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def _create_summary_sheet(wb: Workbook, students: List[dict]):
    """Sheet 1: Summary — Roll No, Total Marks, Max, %, Grade, Unattempted, Flagged."""
    ws = wb.active
    ws.title = "Summary"

    headers = ["Roll No", "Total Marks", "Max Marks", "Percentage", "Grade",
               "Unattempted Questions", "Flagged Questions"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    for i, student in enumerate(students, 2):
        ws.cell(row=i, column=1, value=student.get("roll_number", ""))
        ws.cell(row=i, column=2, value=student.get("total_marks_awarded", 0))
        ws.cell(row=i, column=3, value=student.get("total_marks_possible", 0))
        ws.cell(row=i, column=4, value=student.get("percentage", 0))
        ws.cell(row=i, column=5, value=student.get("grade", ""))
        ws.cell(row=i, column=6, value=", ".join(student.get("unattempted_questions", [])))
        flagged = [f.get("question_number", "") for f in student.get("flagged_questions", [])]
        ws.cell(row=i, column=7, value=", ".join(flagged))
        _apply_cell_border(ws, i, len(headers))

        # Color coding
        pct = student.get("percentage", 0)
        if pct < 40:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = FAIL_FILL
        elif pct >= 70:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = PASS_FILL

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _create_breakdown_sheet(wb: Workbook, students: List[dict]):
    """Sheet 2: Question Breakdown — Roll No + one column pair per question."""
    ws = wb.create_sheet("Question Breakdown")

    # Collect all unique question numbers
    all_questions = set()
    for student in students:
        for q in student.get("question_breakdown", []):
            all_questions.add(q.get("question_number", ""))
    all_questions = sorted(all_questions)

    # Headers
    ws.cell(row=1, column=1, value="Roll No")
    col = 2
    for qnum in all_questions:
        ws.cell(row=1, column=col, value=f"Q{qnum} Awarded")
        ws.cell(row=1, column=col + 1, value=f"Q{qnum} Max")
        col += 2
    _apply_header_style(ws, 1, 1 + len(all_questions) * 2)

    # Data
    for i, student in enumerate(students, 2):
        ws.cell(row=i, column=1, value=student.get("roll_number", ""))
        q_map = {q.get("question_number"): q for q in student.get("question_breakdown", [])}
        col = 2
        for qnum in all_questions:
            q = q_map.get(qnum, {})
            ws.cell(row=i, column=col, value=q.get("marks_awarded", 0))
            ws.cell(row=i, column=col + 1, value=q.get("marks_total", 0))
            col += 2
        _apply_cell_border(ws, i, 1 + len(all_questions) * 2)

    ws.column_dimensions["A"].width = 16
    for col in range(2, 2 + len(all_questions) * 2):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _create_feedback_sheet(wb: Workbook, students: List[dict]):
    """Sheet 3: Detailed Feedback — per question for each student."""
    ws = wb.create_sheet("Detailed Feedback")

    headers = ["Roll No", "Q No", "Type", "Awarded", "Max", "Reasoning",
               "Student Feedback", "Keywords Matched", "Keywords Missing", "Flag"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    row = 2
    for student in students:
        roll = student.get("roll_number", "")
        for q in student.get("question_breakdown", []):
            ws.cell(row=row, column=1, value=roll)
            ws.cell(row=row, column=2, value=q.get("question_number", ""))
            ws.cell(row=row, column=3, value=q.get("question_type", ""))
            ws.cell(row=row, column=4, value=q.get("marks_awarded", 0))
            ws.cell(row=row, column=5, value=q.get("marks_total", 0))
            ws.cell(row=row, column=6, value=q.get("reasoning", ""))
            ws.cell(row=row, column=7, value=q.get("student_feedback", ""))
            ws.cell(row=row, column=8, value=", ".join(q.get("keywords_matched", [])))
            ws.cell(row=row, column=9, value=", ".join(q.get("keywords_missing", [])))
            ws.cell(row=row, column=10, value=q.get("flag", ""))
            _apply_cell_border(ws, row, len(headers))

            if q.get("flag"):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = FLAG_FILL
            row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 40


def _create_statistics_sheet(wb: Workbook, students: List[dict]):
    """Sheet 4: Statistics — class stats, pass rate, grade distribution, per-question avg."""
    ws = wb.create_sheet("Statistics")

    percentages = [s.get("percentage", 0) for s in students]
    grades = [s.get("grade", "") for s in students]

    # Class statistics
    ws.cell(row=1, column=1, value="Class Statistics")
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    stats = [
        ("Total Students", len(students)),
        ("Average %", round(statistics.mean(percentages), 2) if percentages else 0),
        ("Median %", round(statistics.median(percentages), 2) if percentages else 0),
        ("Highest %", max(percentages) if percentages else 0),
        ("Lowest %", min(percentages) if percentages else 0),
        ("Std Deviation", round(statistics.stdev(percentages), 2) if len(percentages) > 1 else 0),
        ("Pass Rate (>=40%)", f"{sum(1 for p in percentages if p >= 40)}/{len(percentages)}"),
    ]

    for i, (label, value) in enumerate(stats, 2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    # Grade distribution
    row = len(stats) + 4
    ws.cell(row=row, column=1, value="Grade Distribution")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    row += 1
    grade_dist = {}
    for g in grades:
        grade_dist[g] = grade_dist.get(g, 0) + 1
    for grade in sorted(grade_dist.keys()):
        ws.cell(row=row, column=1, value=grade).font = Font(bold=True)
        ws.cell(row=row, column=2, value=grade_dist[grade])
        row += 1

    # Per-question average
    row += 2
    ws.cell(row=row, column=1, value="Per-Question Average Marks")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    row += 1

    question_marks: Dict[str, List[float]] = {}
    question_totals: Dict[str, float] = {}
    for student in students:
        for q in student.get("question_breakdown", []):
            qnum = q.get("question_number", "")
            if qnum:
                if qnum not in question_marks:
                    question_marks[qnum] = []
                    question_totals[qnum] = q.get("marks_total", 0)
                question_marks[qnum].append(q.get("marks_awarded", 0))

    ws.cell(row=row, column=1, value="Question")
    ws.cell(row=row, column=2, value="Avg Marks")
    ws.cell(row=row, column=3, value="Max Marks")
    ws.cell(row=row, column=4, value="Avg %")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4).font = Font(bold=True)
    row += 1

    # Sort by average ascending (weakest first)
    sorted_questions = sorted(
        question_marks.items(),
        key=lambda x: statistics.mean(x[1]) if x[1] else 0,
    )

    for qnum, marks_list in sorted_questions:
        avg = round(statistics.mean(marks_list), 2) if marks_list else 0
        total = question_totals.get(qnum, 0)
        ws.cell(row=row, column=1, value=f"Q{qnum}")
        ws.cell(row=row, column=2, value=avg)
        ws.cell(row=row, column=3, value=total)
        ws.cell(row=row, column=4, value=round((avg / total * 100) if total else 0, 1))
        row += 1

    # Unattempted frequency
    row += 2
    ws.cell(row=row, column=1, value="Unattempted Frequency per Question")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    row += 1

    unattempted_freq: Dict[str, int] = {}
    for student in students:
        for qnum in student.get("unattempted_questions", []):
            unattempted_freq[qnum] = unattempted_freq.get(qnum, 0) + 1

    ws.cell(row=row, column=1, value="Question").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Unattempted Count").font = Font(bold=True)
    row += 1
    for qnum in sorted(unattempted_freq.keys()):
        ws.cell(row=row, column=1, value=f"Q{qnum}")
        ws.cell(row=row, column=2, value=unattempted_freq[qnum])
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14


def generate_course_excel(
    course_code: str,
    students: List[dict],
    course_data: Optional[dict],
    output_dir: str,
) -> str:
    """Generate a complete Excel file for a course. Returns the file path."""
    wb = Workbook()

    _create_summary_sheet(wb, students)
    _create_breakdown_sheet(wb, students)
    _create_feedback_sheet(wb, students)
    _create_statistics_sheet(wb, students)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{course_code}_Results_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    logger.info(f"Excel exported: {filepath}")

    return filepath


def generate_all_courses_zip(
    all_results: List[dict],
    all_courses: List[dict],
    output_dir: str,
) -> str:
    """Generate Excel files for all courses and zip them together."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Group results by course
    by_course: Dict[str, List[dict]] = {}
    for r in all_results:
        code = r.get("course_code", "unknown")
        if code not in by_course:
            by_course[code] = []
        by_course[code].append(r)

    course_data_map = {c.get("course_code"): c for c in all_courses}

    excel_files = []
    for course_code, students in by_course.items():
        filepath = generate_course_excel(
            course_code, students, course_data_map.get(course_code), output_dir,
        )
        excel_files.append(filepath)

    # Create zip
    zip_filename = f"All_Results_{timestamp}.zip"
    zip_filepath = os.path.join(output_dir, zip_filename)

    with zipfile.ZipFile(zip_filepath, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in excel_files:
            zf.write(fp, os.path.basename(fp))

    # Remove individual files
    for fp in excel_files:
        try:
            os.remove(fp)
        except Exception:
            pass

    logger.info(f"Zip exported: {zip_filepath}")
    return zip_filepath
